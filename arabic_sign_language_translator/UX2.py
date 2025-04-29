import streamlit as st
import cv2
import keras
import numpy as np
import tensorflow as tf
from openpyxl import load_workbook
from collections import deque
from PIL import Image

# Custom CSS for modern styling
st.markdown("""
    <style>
    .main {
        background-color: #f8f9fa;
    }
    .stButton>button {
        background-color: #4CAF50;
        color: white;
        border-radius: 8px;
        padding: 10px 24px;
        font-weight: bold;
        border: none;
        transition: all 0.3s;
    }
    .stButton>button:hover {
        background-color: #45a049;
        transform: scale(1.05);
    }
    .stop-button>button {
        background-color: #f44336 !important;
    }
    .header {
        color: #2c3e50;
        text-align: center;
        margin-bottom: 30px;
    }
    .prediction-box {
        background-color: white;
        border-radius: 12px;
        padding: 20px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        margin-top: 20px;
    }
    .prediction-text {
        font-size: 24px;
        font-weight: bold;
        color: #2c3e50;
    }
    .confidence-text {
        font-size: 18px;
        color: #7f8c8d;
    }
    .video-container {
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 4px 12px rgba(0,0,0,0.15);
        margin-bottom: 20px;
    }
    </style>
    """, unsafe_allow_html=True)

# Load model
@st.cache_resource
def load_model():
    return keras.models.load_model('100_labels_model_128x128_30_frames.keras')

model = load_model()

# Load label map
@st.cache_resource
def load_label_map(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    sign_ids = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        sign_id = str(row[0]).zfill(4)
        sign_ids.append(sign_id)

    label_map = {sid: idx for idx, sid in enumerate(sign_ids)}
    signid_map = {idx: sid for idx, sid in enumerate(sign_ids)}
    return label_map, signid_map

label_map, signid_map = load_label_map("KARSL-100_Labels.xlsx")

# Settings
input_size = (128, 128)
max_frames = 30
predict_every_n_frames = 60

# App layout
st.markdown('<h1 class="header">Sign Language Recognition 🤟</h1>', unsafe_allow_html=True)
st.markdown("""
    <div style="text-align: center; margin-bottom: 30px; color: #7f8c8d;">
    Real-time sign language detection using deep learning.
    Start the camera and show your signs to the webcam.
    </div>
    """, unsafe_allow_html=True)

# Session state
if 'camera_running' not in st.session_state:
    st.session_state.camera_running = False
    st.session_state.prediction = None
    st.session_state.confidence = None
    st.session_state.frame_buffer = deque(maxlen=max_frames)

# Sidebar for settings
with st.sidebar:
    st.header("Settings ⚙️")
    predict_every_n_frames = st.slider("Prediction frequency (frames)", 1, 120, 30)
    st.markdown("---")
    st.markdown("""
        **How to use:**
        1. Click 'Start Camera'
        2. Show your sign to the webcam
        3. View predictions in real-time
        4. Click 'Stop Camera' when done
        """)

# Button layout
col1, col2 = st.columns(2)
with col1:
    if st.button('🎥 Start Camera', key='start'):
        st.session_state.camera_running = True
        st.session_state.prediction = None
        st.session_state.confidence = None
        st.session_state.frame_buffer = deque(maxlen=max_frames)

with col2:
    if st.button('🛑 Stop Camera', key='stop'):
        st.session_state.camera_running = False

# Video and prediction display
video_placeholder = st.empty()
prediction_placeholder = st.empty()

if st.session_state.camera_running:
    cap = cv2.VideoCapture(0)
    frame_count = 0

    while st.session_state.camera_running:
        ret, frame = cap.read()
        if not ret:
            st.error("Failed to access camera. Please check your webcam connection.")
            break

        # Preprocess
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        resized_frame = cv2.resize(frame_rgb, input_size)
        normalized_frame = resized_frame / 255.0
        st.session_state.frame_buffer.append(normalized_frame)

        # Prediction
        if len(st.session_state.frame_buffer) == max_frames and frame_count % predict_every_n_frames == 0:
            input_batch = np.expand_dims(np.array(st.session_state.frame_buffer), axis=0)
            preds = model.predict(input_batch, verbose=0)
            pred_class = np.argmax(preds)
            st.session_state.prediction = signid_map[pred_class]
            st.session_state.confidence = np.max(preds)

        # Display prediction on frame
        if st.session_state.prediction:
            cv2.putText(frame, f"{st.session_state.prediction}", (20, 50),
                        cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 200, 0), 3)
            cv2.putText(frame, f"Confidence: {st.session_state.confidence:.2f}", (20, 100),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 180, 0), 2)

        # Display in Streamlit with modern styling
        with video_placeholder.container():
            st.markdown('<div class="video-container">', unsafe_allow_html=True)
            st.image(frame, channels="BGR", use_column_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        # Display prediction box
        if st.session_state.prediction:
            with prediction_placeholder.container():
                st.markdown('<div class="prediction-box">', unsafe_allow_html=True)
                st.markdown(f'<p class="prediction-text">Predicted Sign: {st.session_state.prediction}</p>', unsafe_allow_html=True)
                st.markdown(f'<p class="confidence-text">Confidence: {st.session_state.confidence:.2%}</p>', unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)

        frame_count += 1

    cap.release()
else:
    # Display placeholder when camera is off
    with video_placeholder.container():
        st.markdown('<div class="video-container">', unsafe_allow_html=True)
        st.image(Image.new('RGB', (640, 480), color=(240, 240, 240)))
        st.markdown('</div>', unsafe_allow_html=True)

    if st.session_state.prediction is None:
        with prediction_placeholder.container():
            st.markdown('<div class="prediction-box">', unsafe_allow_html=True)
            st.markdown('<p class="prediction-text">Camera is off. Start the camera to begin recognition.</p>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

# Footer
st.markdown("""
    <div style="text-align: center; margin-top: 50px; color: #95a5a6; font-size: 14px;">
    Made with ❤️ using Streamlit | Sign Language Recognition System
    </div>
    """, unsafe_allow_html=True)
