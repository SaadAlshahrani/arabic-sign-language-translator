import streamlit as st
import cv2
import keras
import numpy as np
from openpyxl import load_workbook
from collections import deque
from PIL import Image
from streamlit_TTS import auto_play, text_to_speech, text_to_audio
from gtts.lang import tts_langs

langs = tts_langs().keys()

#---------------------------Set page configuration------------------------------
st.set_page_config(
    page_title="ASLT",
    page_icon="🤟",
    layout="wide",
    initial_sidebar_state="expanded",
)

#---------------------------Custom CSS for styling------------------------------
st.markdown(
    """
    <style>
    .main-title {
        font-size: 36px;
        font-weight: bold;
        color: #4CAF50;
    }
    .Main-Title {
            font-size: 58px;
            font-weight: bold;
            text-align: center;
            background: -webkit-linear-gradient(45deg, #004474, #ff5733, #28a745);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }
    .SideBar {
        font-size: 58px;
            font-weight: bold;
            text-align: center;
            background: -webkit-linear-gradient(45deg, #004474, #ff5733, #28a745);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <style>
    .start_camera {
        background-color: #4CAF50;
        color: white;
        font-size: 18px;
        font-weight: bold;
        padding: 10px 20px;
        border: none;
        border-radius: 5px;
        cursor: pointer;
        margin-right: 10px;
    }
    .start_camera:hover {
        background-color: #4CAF50;
        color: white;
    }
    .stop_camera {
        background-color: #f44336;
        color: white;
        font-size: 18px;
        font-weight: bold;
        padding: 10px 20px;
        border: none;
        border-radius: 5px;
        cursor: pointer;
    }
    .stop_camera:hover {
        background-color: #e53935;
    }
    .prediction-text{
        font-size: 80px;
        font-weight: bold;
        background: -webkit-linear-gradient(45deg, #004474, #ff5733, #28a745);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .confidence-text{
        font-size: 80px;
        font-weight: bold;
        background: -webkit-linear-gradient(45deg, #004474, #ff5733, #28a745);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    .Cam-OFF-Text{
        font-size: 80px;
        font-weight: bold;
        background: -webkit-linear-gradient(45deg, #004474, #ff5733, #28a745);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

#-----------------------------------Sidebar-------------------------------------

with st.sidebar.title("ASLT"):
    st.markdown(
        """
        <h1 class='SideBar' style='color:orange; font-size:70px;'>ASLT</h1>
        """,
        unsafe_allow_html=True,
    )
    st.sidebar.markdown("---")

# st.sidebar.title("Navigation")
# page = st.sidebar.radio("Go to", ["🏠 Home", "🈲 Translator", "⚙️ Settings"])

# if page == "🏠 Home":
#     from Pages import Home
#     Home.app()
# elif page == "🈲 Translator":
#     from Pages import translator
#     translator.app()
# elif page == "⚙️ Settings":
#     from Pages import settings
#     settings.app()

with st.sidebar.expander("Settings ⚙️", expanded=True):
    st.header("Edit The Frames")
    predict_every_n_frames = st.slider("Prediction frequency ", 1, 120, 30)
    st.sidebar.markdown("---")

with st.sidebar.expander("Choose Language 🪄", expanded=True):
    lang = st.selectbox("", options=langs)
    speak = st.button("Speak it out!")
    st.sidebar.markdown("---")

with st.sidebar.expander("Text to Speech", expanded=True):
    text = st.text_input("Choose a text to speak out:")
    st.sidebar.markdown("---")

    st.sidebar.markdown("""
        <div style="font-size: 18px; line-height: 1.6; margin-bottom: 20px;">
        <strong>How to use ASLT:</strong><br>
        1. Click <strong>'Start Camera'</strong><br>
        2. Show your sign to the webcam<br>
        3. View predictions in real-time<br>
        4. Click <strong>'Stop Camera'</strong> when done
        </div>
        """, unsafe_allow_html=True)

#--------------------------------Main page--------------------------------------
st.markdown("<div class='Main-Title'>Arabic Sign Language Translator </div>" , unsafe_allow_html=True)
st.markdown("---")

st.markdown('<p class="">Welcome to the Arabic Sign Language Translator 🤟🏻🕵🏻‍♂️!</p>', unsafe_allow_html=True)


#-----------------------------------Camera Control--------------------------------
video_placeholder = st.empty()
prediction_placeholder = st.empty()

# Load model
@st.cache_resource
def load_model():
    return keras.models.load_model('100_labels_model_128x128_30_frames.keras')

model = load_model()

# Load label map
@st.cache_resource
def load_label_map(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    sign_ids = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        sign_id = str(row[1]).zfill(4)
        sign_ids.append(sign_id)

    label_map = {sid: idx for idx, sid in enumerate(sign_ids)}
    signid_map = {idx: sid for idx, sid in enumerate(sign_ids)}
    return label_map, signid_map

label_map, signid_map = load_label_map("KARSL-100_Labels.xlsx")

# Settings
input_size = (128, 128)
max_frames = 30
predict_every_n_frames = 60


# Session state
if 'camera_running' not in st.session_state:
    st.session_state.camera_running = False
    st.session_state.prediction = None
    st.session_state.confidence = None
    st.session_state.frame_buffer = deque(maxlen=max_frames)

col1, col2 = st.columns(2)
with col1:
    if st.button("🎥🟢 Start Camera", key="start_camera", use_container_width=True):
        st.session_state.camera_running = True
        st.session_state.prediction = None
        st.session_state.confidence = None
        st.session_state.frame_buffer = deque(maxlen=max_frames)
        st.write("Camera is ON")

with col2:
    if st.button("🎥🛑 Stop Camera", key="stop_camera", use_container_width=True):
        st.session_state.camera_running = False
        st.write("Camera is OFF")


def speak_prediction(prediction, lang='en'):
    text_to_speech(text=prediction, language=lang)

# Main logic to process webcam feed and predictions
video_placeholder = st.empty()
prediction_placeholder = st.empty()

if st.session_state.camera_running:
    cap = cv2.VideoCapture(0)
    frame_count = 0
    while st.session_state.camera_running:
        ret, frame = cap.read()
        if not ret:
            st.error("Failed to access camera. Please check your webcam connection.")
            break

        # Preprocess
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        resized_frame = cv2.resize(frame_rgb, input_size)
        normalized_frame = resized_frame / 255.0
        st.session_state.frame_buffer.append(normalized_frame)

        # Prediction
        if len(st.session_state.frame_buffer) == max_frames and frame_count % predict_every_n_frames == 0:
            input_batch = np.expand_dims(np.array(st.session_state.frame_buffer), axis=0)
            preds = model.predict(input_batch, verbose=0)
            pred_class = np.argmax(preds)
            st.session_state.prediction = signid_map[pred_class]
            st.session_state.confidence = np.max(preds)

        # Display prediction on frame
        if st.session_state.prediction:
            cv2.putText(frame, f"{st.session_state.prediction}", (20, 50),
                        cv2.FONT_HERSHEY_SIMPLEX, 1.2, (0, 200, 0), 3)
            cv2.putText(frame, f"Confidence: {st.session_state.confidence:.2f}", (20, 100),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 180, 0), 2)

            # Speak the prediction if it's a new one
            speak_prediction(st.session_state.prediction)

        # Display in Streamlit with modern styling
        with video_placeholder.container():
            st.markdown('<div class="video-container">', unsafe_allow_html=True)
            st.image(frame, channels="BGR", use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)

        # Display prediction box
        if st.session_state.prediction:
            with prediction_placeholder.container():
                st.markdown('<div class="prediction-box">', unsafe_allow_html=True)
                st.markdown(f'<p class="prediction-text" style="font-size: 50px;">Predicted Sign: <strong> {st.session_state.prediction} </strong></p>', unsafe_allow_html=True)
                st.markdown(f'<p class="confidence-text" style="font-size: 50px;">Confidence:<strong> {st.session_state.confidence:.2%}</strong></p>', unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)
        frame_count += 1
    cap.release()
else:
    if st.session_state.prediction is None:
        with prediction_placeholder.container():
            st.markdown('<div class="prediction-box">', unsafe_allow_html=True)
            st.markdown('<p class="Cam-OFF-Text" style="font-size: 40px;">Camera is off. Start the camera to begin recognition.</p>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
    if st.session_state.prediction is None:
        with prediction_placeholder.container():
            st.markdown('<div class="prediction-box">', unsafe_allow_html=True)
            st.markdown('<p class="Cam-OFF-Text" style="font-size: 40px;">Camera is off. Start the camera to begin recognition.</p>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
#-----------------------------------Model Loading-------------------------------



















#-----------------------------------About Section-------------------------------
st.write(
    """
    This application is designed to help translate Arabic sign language into text or speech.
    It leverages advanced machine learning models to provide accurate and real-time translations.
    """
)

# Features section
st.header("Features")
st.write(
    """
    - **Real-time Translation**: Translate Arabic sign language gestures into text or speech instantly.
    - **User-Friendly Interface**: Simple and intuitive design for ease of use.
    - **Customizable Options**: Adjust settings to suit your preferences.
    """
)

# Call to action
st.header("Get Started")
st.write("Click on the **Translator** tab in the sidebar to begin translating Arabic sign language.")

# Footer
st.markdown("---")
st.markdown(
    """
    <footer style="text-align: center;">
        © 2025 Arabic Sign Language Translator. All rights reserved.
    </footer>
    """,
    unsafe_allow_html=True,
)
