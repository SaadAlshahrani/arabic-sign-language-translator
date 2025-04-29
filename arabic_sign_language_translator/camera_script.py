import cv2
import keras
import numpy as np
import tensorflow as tf
import time
from openpyxl import load_workbook
from collections import deque

def load_label_map(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    sign_ids = [str(row[0]).zfill(4) for row in ws.iter_rows(min_row=2, values_only=True)]
    label_map = {sid: idx for idx, sid in enumerate(sign_ids)}
    signid_map = {idx: sid for idx, sid in enumerate(sign_ids)}
    return label_map, signid_map

def preprocess_frame(frame, input_size):
    frame = cv2.resize(frame, input_size)
    return frame / 255.0

def detect_motion(prev_gray, frame, threshold=500000):
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    if prev_gray is None:
        return gray, False, 0
    diff = cv2.absdiff(prev_gray, gray)
    motion_score = np.sum(diff)
    return gray, motion_score > threshold, motion_score

def predict_sign(model, buffer, signid_map):
    input_batch = np.expand_dims(np.array(buffer), axis=0)
    preds = model.predict(input_batch, verbose=0)
    pred_class = np.argmax(preds)
    confidence = np.max(preds)
    return signid_map[pred_class], confidence

def main_script():
    # Load model and label map
    model = keras.models.load_model('100_labels_model_128x128_30_frames_lstm.keras')
    label_map, signid_map = load_label_map("KARSL-100_Labels.xlsx")

    # Parameters
    input_size = (128, 128)
    max_frames = 30
    capture_every_n_frames = 2
    pause_duration = 3.0  # seconds
    motion_threshold = 500000

    # State variables
    frame_buffer = deque(maxlen=max_frames)
    frame_count = 0
    last_pred_label = ""
    last_confidence = 0.0
    paused = False
    last_prediction_time = 0
    prev_gray = None

    cap = cv2.VideoCapture(0)

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        current_time = time.time()
        status_text = ""
        display_prediction = False

        if paused:
            if current_time - last_prediction_time >= pause_duration:
                paused = False
            else:
                status_text = "⏸ Paused - Showing Prediction"
                display_prediction = True
        else:
            prev_gray, motion_detected, motion_score = detect_motion(prev_gray, frame, motion_threshold)

            if motion_detected:
                status_text = "📸 Taking frames..."
                frame_count += 1
                if frame_count % capture_every_n_frames == 0:
                    processed = preprocess_frame(frame, input_size)
                    frame_buffer.append(processed)
                    if len(frame_buffer) == max_frames:
                        last_pred_label, last_confidence = predict_sign(model, frame_buffer, signid_map)
                        last_prediction_time = current_time
                        paused = True
                        frame_buffer.clear()
                        frame_count = 0
            else:
                status_text = "🤚 Waiting for motion..."

        # Display status
        cv2.putText(frame, status_text, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 0), 2)

        # Display prediction
        if last_pred_label and (display_prediction or not paused):
            cv2.putText(frame, f"🧠 {last_pred_label} ({last_confidence:.2f})", (10, 65),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

        # Display progress bar
        if not paused and "Taking frames" in status_text:
            bar_x, bar_y = 10, frame.shape[0] - 30
            bar_width = frame.shape[1] - 20
            progress = min(len(frame_buffer) / max_frames, 1.0)
            filled = int(bar_width * progress)
            cv2.rectangle(frame, (bar_x, bar_y), (bar_x + bar_width, bar_y + 20), (50, 50, 50), -1)  # background
            cv2.rectangle(frame, (bar_x, bar_y), (bar_x + filled, bar_y + 20), (0, 255, 0), -1)      # filled
            cv2.rectangle(frame, (bar_x, bar_y), (bar_x + bar_width, bar_y + 20), (255, 255, 255), 2)  # border

        cv2.imshow('Sign Language Prediction', frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    main_script()
