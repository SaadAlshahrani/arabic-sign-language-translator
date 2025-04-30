import cv2
import keras

import numpy as np
import tensorflow as tf

from openpyxl import load_workbook
from collections import deque


model = keras.models.load_model('100_labels_model_128x128_30_frames.keras')  # 🔥 Update this to your model path

def load_label_map(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    sign_ids = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        sign_id = str(row[0]).zfill(4)  # Assuming 'SignID' is the first column
        sign_ids.append(sign_id)

    label_map = {sid: idx for idx, sid in enumerate(sign_ids)}
    signid_map = {idx: sid for idx, sid in enumerate(sign_ids)}

    return label_map, signid_map

# Usage
label_map, signid_map = load_label_map("KARSL-100_Labels.xlsx")

# Parameters
input_size = (128, 128)  # Match your model input size
max_frames = 30        # Number of frames the model expects
frame_count = 0
predict_every_n_frames = 15

# Initialize video capture
cap = cv2.VideoCapture(0)

# Frame buffer
frame_buffer = deque(maxlen=max_frames)

def preprocess_frame(frame):
    frame = cv2.resize(frame, input_size)
    frame = frame / 255.0  # Normalize
    return frame

while True:
    ret, frame = cap.read()
    if not ret:
        break

    # Preprocess and add to buffer
    processed_frame = preprocess_frame(frame)
    frame_buffer.append(processed_frame)

    # Only predict if we have enough frames
    if len(frame_buffer) == max_frames and frame_count % predict_every_n_frames == 0:
        input_batch = np.array(frame_buffer)
        input_batch = np.expand_dims(input_batch, axis=0)  # Add batch dimension

        # Predict
        preds = model.predict(input_batch, verbose=0)
        pred_class = np.argmax(preds)
        pred_label = signid_map[pred_class]
        print(pred_label)
        confidence = np.max(preds)

        # Display prediction
        cv2.putText(frame, f"{pred_label} ({confidence:.2f})", (10, 30),
                    cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)

    # Show the frame
    cv2.imshow('Sign Language Prediction', frame)

    # Exit if 'q' is pressed
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release resources
cap.release()
cv2.destroyAllWindows()
